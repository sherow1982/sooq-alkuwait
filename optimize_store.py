import os
import re
import random
import json
from urllib.parse import quote
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ================= إعدادات المتجر =================
BASE_URL = "https://sooq-alkuwait.arabsad.com"
PRODUCTS_DIR = "products-pages"
FEED_FILENAME = "google_dsa_feed_final.xlsx"

# ================= القوالب (Templates) =================
INTRO_TEMPLATES = [
    "هل تبحث عن <strong>{name}</strong> بسعر مميز؟ في سوق الكويت، نقدم لك أفضل العروض مع ضمان الجودة.",
    "احصل الآن على <strong>{name}</strong> الأصلي بأفضل سعر في الكويت. تجربة تسوق سهلة ومضمونة.",
    "لا تفوت فرصة اقتناء <strong>{name}</strong>. نوفر لك منتجات عالية الجودة بأسعار تنافسية لا تقبل المقارنة.",
    "تسوق <strong>{name}</strong> الآن واستمتع برفاهية التسوق من المنزل مع سوق الكويت. خيارك الأول للجودة.",
    "إذا كنت تريد <strong>{name}</strong> بمواصفات ممتازة وسعر رائع، فأنت في المكان الصحيح."
]

DELIVERY_TEMPLATES = [
    "نتميز بخدمة توصيل سريعة تغطي جميع مناطق الكويت (العاصمة، الجهراء، الفروانية، حولي، مبارك الكبير، والأحمدي).",
    "يصلك طلبك لباب المنزل في أي مكان بالكويت خلال 24-48 ساعة. تغطية شاملة لجميع المحافظات.",
    "شحن سريع وموثوق لجميع مناطق الكويت. اطلب اليوم واستلم طلبك في أسرع وقت.",
    "خدمة التوصيل لدينا تغطي كافة أنحاء الكويت (من الجهراء إلى الأحمدي) بسرعة وكفاءة عالية."
]

CONCLUSION_TEMPLATES = [
    "اطلب <strong>{name}</strong> الآن واستفد من عروضنا الحصرية لفترة محدودة. الكمية متوفرة!",
    "سارع بطلب <strong>{name}</strong> قبل نفاذ الكمية. أفضل الأسعار بانتظارك.",
    "أضف <strong>{name}</strong> إلى سلتك اليوم واستمتع بتجربة تسوق لا مثيل لها.",
    "لا تتردد، <strong>{name}</strong> هو الخيار الأمثل لك. اطلبه الآن عبر الموقع أو الواتساب."
]

def get_clean_schema(product_name, sku, url, price, delivery_text):
    clean_price = re.sub(r'[^\d.]', '', str(price))
    schema_data = {
        "@context": "https://schema.org",
        "@type": "Product",
        "name": product_name,
        "description": f"اشتري {product_name} اونلاين في الكويت. {delivery_text}",
        "sku": sku,
        "mpn": sku,
        "brand": {
            "@type": "Brand",
            "name": "سوق الكويت"
        },
        "offers": {
            "@type": "Offer",
            "url": url,
            "priceCurrency": "KWD",
            "price": clean_price,
            "priceValidUntil": "2026-12-31",
            "availability": "https://schema.org/InStock",
            "itemCondition": "https://schema.org/NewCondition",
            "shippingDetails": {
                "@type": "OfferShippingDetails",
                "shippingRate": {
                    "@type": "MonetaryAmount",
                    "value": "0.0",
                    "currency": "KWD"
                },
                "shippingDestination": {
                    "@type": "DefinedRegion",
                    "addressCountry": "KW"
                },
                "deliveryTime": {
                    "@type": "ShippingDeliveryTime",
                    "handlingTime": {
                        "@type": "QuantitativeValue",
                        "minValue": 1,
                        "maxValue": 2,
                        "unitCode": "DAY"
                    },
                    "transitTime": {
                        "@type": "QuantitativeValue",
                        "minValue": 1,
                        "maxValue": 3,
                        "unitCode": "DAY"
                    }
                }
            },
            "hasMerchantReturnPolicy": {
                "@type": "MerchantReturnPolicy",
                "returnPolicyCategory": "https://schema.org/MerchantReturnFiniteReturnWindow",
                "merchantReturnDays": 14,
                "returnMethod": "https://schema.org/ReturnByMail",
                "returnFees": "https://schema.org/FreeReturn"
            }
        }
    }
    return json.dumps(schema_data, ensure_ascii=False, indent=2)

def generate_seo_block(product_name):
    intro = random.choice(INTRO_TEMPLATES).format(name=product_name)
    delivery = random.choice(DELIVERY_TEMPLATES)
    conclusion = random.choice(CONCLUSION_TEMPLATES).format(name=product_name)
    
    html = f"""
    <div class="seo-content" style="margin: 30px auto; padding: 25px; background: #f8f9fa; border-radius: 12px; border: 1px solid #e9ecef; max-width: 1200px; text-align: right; direction: rtl;">
        <h2 style="color: #2c3e50; font-size: 1.5rem; margin-bottom: 15px; border-bottom: 2px solid #3498db; padding-bottom: 10px; display: inline-block;">تسوق {product_name} في الكويت</h2>
        <p style="color: #555; line-height: 1.8; font-size: 1.1rem; margin-bottom: 20px;">
            {intro} {delivery}
        </p>
        
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin-bottom: 20px;">
            <div style="background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05);">
                <h3 style="color: #e67e22; margin-bottom: 10px; font-size: 1.1rem;">📦 توصيل سريع ومجاني</h3>
                <p style="font-size: 0.95rem; color: #666;">خدمة توصيل مميزة لجميع المناطق.</p>
            </div>
            <div style="background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05);">
                <h3 style="color: #27ae60; margin-bottom: 10px; font-size: 1.1rem;">✅ جودة مضمونة</h3>
                <p style="font-size: 0.95rem; color: #666;">منتجات أصلية ومفحوصة بعناية.</p>
            </div>
            <div style="background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05);">
                <h3 style="color: #8e44ad; margin-bottom: 10px; font-size: 1.1rem;">💳 دفع آمن وسهل</h3>
                <p style="font-size: 0.95rem; color: #666;">خيارات دفع متعددة وآمنة 100%.</p>
            </div>
        </div>

        <p style="color: #777; font-size: 0.9rem; margin-top: 20px;">
            {conclusion}
        </p>
    </div>
    """
    return html

def clean_product_name(filename):
    name = filename.replace('.html', '')
    name = re.sub(r'^product-\d+-', '', name)
    name = name.replace('-', ' ')
    return name.strip()

def extract_price(soup):
    price_element = soup.find(string=re.compile(r'\d+(\.\d+)?\s*(KWD|د\.ك)'))
    if price_element:
        match = re.search(r'(\d+(\.\d+)?)', price_element)
        if match:
            return match.group(1)
    return "10.0"

def optimize_files():
    feed_data = []
    
    if not os.path.exists(PRODUCTS_DIR):
        print(f"❌ المجلد {PRODUCTS_DIR} غير موجود!")
        return

    files = [f for f in os.listdir(PRODUCTS_DIR) if f.endswith('.html')]
    print(f"🚀 جاري معالجة {len(files)} منتج...")

    for filename in files:
        filepath = os.path.join(PRODUCTS_DIR, filename)
        product_name = clean_product_name(filename)
        encoded_filename = quote(filename)
        product_url = f"{BASE_URL}/{PRODUCTS_DIR}/{encoded_filename}"
        
        with open(filepath, 'r', encoding='utf-8') as f:
            original_html = f.read()
            soup = BeautifulSoup(original_html, 'html.parser')

        current_price = extract_price(soup)
        delivery_text = random.choice(DELIVERY_TEMPLATES)
        sku = filename.replace('.html', '')

        # تحديث الميتا والعناوين
        if soup.head:
            if soup.title: soup.title.decompose()
            new_title = soup.new_tag('title')
            new_title.string = f"{product_name} | أفضل سعر في الكويت"
            soup.head.append(new_title)
            
            old_meta = soup.find('meta', attrs={'name': 'description'})
            if old_meta: old_meta.decompose()
            meta_desc_text = f"اشتري {product_name} اونلاين بأفضل سعر في الكويت. {delivery_text} تسوق الآن!"
            new_meta = soup.new_tag('meta', attrs={'name': 'description', 'content': meta_desc_text})
            soup.head.append(new_meta)

            old_canonical = soup.find('link', attrs={'rel': 'canonical'})
            if old_canonical: old_canonical.decompose()
            new_canonical = soup.new_tag('link', attrs={'rel': 'canonical', 'href': product_url})
            soup.head.append(new_canonical)

        # تحديث السكيما
        old_schemas = soup.find_all('script', type='application/ld+json')
        for s in old_schemas:
            s.decompose()

        clean_json_string = get_clean_schema(product_name, sku, product_url, current_price, delivery_text)
        new_schema_tag = soup.new_tag('script', type='application/ld+json')
        new_schema_tag.string = clean_json_string
        
        if soup.head:
            soup.head.append(new_schema_tag)

        # تحديث بلوك الـ SEO
        old_seo_content = soup.find('div', class_='seo-content')
        if old_seo_content:
            old_seo_content.decompose()

        seo_html = generate_seo_block(product_name)
        seo_tag = BeautifulSoup(seo_html, 'html.parser')
        
        if soup.body:
            footer = soup.find('footer')
            if footer:
                footer.insert_before(seo_tag)
            else:
                soup.body.append(seo_tag)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(str(soup))

        # جمع البيانات للملف
        feed_data.append([product_url, "BestSeller"])

    # 🔥 إنشاء ملف XLSX (Excel حقيقي)
    print("📊 جاري إنشاء ملف Excel...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Feed"
    
    # تنسيق الهيدر
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # كتابة الهيدر
    ws['A1'] = "Page URL"
    ws['B1'] = "Custom Label"
    
    # تطبيق التنسيق على الهيدر
    for cell in ['A1', 'B1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
    
    # تعديل عرض الأعمدة
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20
    
    # كتابة البيانات
    for idx, (url, label) in enumerate(feed_data, start=2):
        ws[f'A{idx}'] = url
        ws[f'B{idx}'] = label
    
    # حفظ الملف
    wb.save(FEED_FILENAME)
    
    print(f"✅ تم تحديث {len(files)} صفحة بنجاح!")
    print(f"📊 تم إنشاء ملف الفيد: {FEED_FILENAME}")
    print(f"👉 الملف الآن صيغة XLSX (Excel) بعمودين نظيفين: Page URL و Custom Label")

if __name__ == "__main__":
    optimize_files()
